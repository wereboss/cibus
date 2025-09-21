"""
Cibus: XLS Layout Parser

This module provides a function to parse an XLS file containing the fixed-length
data layout. It is a critical component for the 'Layout Parsing Node' in our
agentic pipeline.
"""

import openpyxl
import os
import re
from typing import List, Tuple, Union

# Regex to find numeric value inside parentheses, e.g., X(10) or 9(12)
COBOL_LENGTH_PATTERN = re.compile(r'\((.+)\)')

def parse_xls_layout(filepath: str) -> List[Tuple[str, int]]:
    """
    Parses an XLS file to extract field names and lengths from a layout.

    The function assumes the following structure in the first worksheet:
    - Column 1 (A): Field Name
    - Column 2 (B): Field Length (can be an integer or a string like "X(10)" or "9(12)")

    Args:
        filepath: The path to the XLS (or XLSX) layout file.

    Returns:
        A list of tuples, where each tuple contains the field name (str)
        and its length (int).

    Raises:
        FileNotFoundError: If the specified file does not exist.
        KeyError: If the worksheet 'Sheet1' is not found.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Layout file not found at: {filepath}")

    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
    except FileNotFoundError:
        raise
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return []

    layout = []
    # Iterate through rows starting from the second row to skip headers
    for row in sheet.iter_rows(min_row=2, values_only=True):
        field_name = row[0]
        field_length = row[1]

        if field_name is None:
            continue

        length_int: Union[int, None] = None
        
        # Check if the field length is a string with a COBOL-style pattern
        if isinstance(field_length, str):
            match = COBOL_LENGTH_PATTERN.search(field_length)
            if match:
                try:
                    length_int = int(match.group(1))
                except ValueError:
                    print(f"Warning: Could not parse length from '{field_length}'. Skipping row.")
                    continue
            else:
                print(f"Warning: Invalid string format for length '{field_length}'. Skipping row.")
                continue
        elif isinstance(field_length, (int, float)):
            length_int = int(field_length)
        else:
            print(f"Skipping row with invalid length type: {field_length}")
            continue

        if length_int is not None:
            layout.append((str(field_name), length_int))

    return layout

if __name__ == '__main__':
    # --- Dummy data and file for isolated testing ---
    print("--- Testing `layout_parser.py` in isolation ---")
    
    # Dummy layout data with COBOL formats
    dummy_layout_data = [
        ["Field Name", "Field Length"],
        ["id", "9(8)"],
        ["product_code", "X(10)"],
        ["product_type", "X(15)"],
        ["price", "9(7)"],
        ["currency", "X(3)"],
        ["status", "X(10)"],
        ["order_date", "9(8)"],
        ["customer_id", "X(12)"],
        ["zip_code", "9(5)"],
        ["quantity", "9(4)"]
    ]

    # Create a dummy XLS file for testing
    dummy_filepath = "dummy_layout.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_data in dummy_layout_data:
        sheet.append(row_data)
    workbook.save(dummy_filepath)
    print(f"Created a dummy layout file at: {dummy_filepath}")

    try:
        # Parse the dummy layout file
        parsed_output = parse_xls_layout(dummy_filepath)
        
        # Print the results for verification
        print("\nParsed Layout (List of Tuples):")
        for item in parsed_output:
            print(item)
            
        # Example of accessing the first parsed item
        if parsed_output:
            print(f"\nExample: First field name is: {parsed_output[0][0]}, with length: {parsed_output[0][1]}")
        else:
            print("\nNo layout parsed.")

    finally:
        # Clean up the dummy file
        if os.path.exists(dummy_filepath):
            os.remove(dummy_filepath)
            print(f"\nCleaned up dummy file: {dummy_filepath}")